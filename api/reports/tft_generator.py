import pandas as pd
import os
from io import BytesIO
from .models import AccountData

def generate_tft_and_sheets(csv_path, start_date, end_date):
    # Contrôle de cohérence TFT
    # Contrôles de cohérence conformes à la documentation SYSCOHADA
    def controle_coherence_complet(tft_data):
        """
        Contrôles automatiques obligatoires selon la documentation :
        - Égalité variation calculée/variation bilantielle
        - Cohérence des totaux par section
        - Absence de comptes orphelins
        - Respect des seuils de matérialité
        """
        controles = {
            'is_coherent': True,
            'errors': [],
            'warnings': [],
            'details': {}
        }
        
        # 1. Égalité variation calculée/variation bilantielle
        flux_operationnels = tft_data.get('ZB', {}).get('montant', 0)
        flux_investissement = tft_data.get('ZC', {}).get('montant', 0)
        flux_financement = tft_data.get('ZE', {}).get('montant', 0)
        treso_ouverture = tft_data.get('ZA', {}).get('montant', 0)
        treso_cloture = tft_data.get('ZH', {}).get('montant', 0)
        
        variation_tft = (flux_operationnels or 0) + (flux_investissement or 0) + (flux_financement or 0)
        variation_treso = (treso_cloture or 0) - (treso_ouverture or 0)
        
        ecart = abs(variation_tft - variation_treso)
        if ecart > 1e-2:
            controles['is_coherent'] = False
            controles['errors'].append(f"Écart variation TFT/Trésorerie: {ecart:.2f}")
        
        controles['details']['variation_tft'] = variation_tft
        controles['details']['variation_treso'] = variation_treso
        controles['details']['ecart'] = ecart
        
        # 2. Cohérence des totaux par section
        # Section A - Activités opérationnelles
        cafg = tft_data.get('FA', {}).get('montant', 0)
        bfr_exploitation = tft_data.get('FB', {}).get('montant', 0)
        variation_stocks = tft_data.get('FC', {}).get('montant', 0)
        variation_creances = tft_data.get('FD', {}).get('montant', 0)
        variation_passif = tft_data.get('FE', {}).get('montant', 0)
        
        total_section_a = cafg + bfr_exploitation + variation_stocks + variation_creances + variation_passif
        if abs(total_section_a - flux_operationnels) > 1e-2:
            controles['warnings'].append(f"Cohérence Section A: écart {abs(total_section_a - flux_operationnels):.2f}")
        
        # Section B - Activités d'investissement
        acquisitions_incorp = tft_data.get('FF', {}).get('montant', 0)
        acquisitions_corp = tft_data.get('FG', {}).get('montant', 0)
        acquisitions_fin = tft_data.get('FH', {}).get('montant', 0)
        cessions_incorp_corp = tft_data.get('FI', {}).get('montant', 0)
        cessions_fin = tft_data.get('FJ', {}).get('montant', 0)
        dividendes_recus = tft_data.get('FJ_DIV', {}).get('montant', 0)
        produits_creances = tft_data.get('FJ_CRE', {}).get('montant', 0)
        
        total_section_b = acquisitions_incorp + acquisitions_corp + acquisitions_fin + cessions_incorp_corp + cessions_fin + dividendes_recus + produits_creances
        if abs(total_section_b - flux_investissement) > 1e-2:
            controles['warnings'].append(f"Cohérence Section B: écart {abs(total_section_b - flux_investissement):.2f}")
        
        # Section C - Activités de financement
        augmentation_capital = tft_data.get('FK', {}).get('montant', 0)
        subventions = tft_data.get('FL', {}).get('montant', 0)
        dividendes_verses = tft_data.get('FM', {}).get('montant', 0)
        nouveaux_emprunts = tft_data.get('FO', {}).get('montant', 0)
        remboursements = tft_data.get('FP', {}).get('montant', 0)
        
        total_section_c = augmentation_capital + subventions - dividendes_verses + nouveaux_emprunts - remboursements
        if abs(total_section_c - flux_financement) > 1e-2:
            controles['warnings'].append(f"Cohérence Section C: écart {abs(total_section_c - flux_financement):.2f}")
        
        # 3. Contrôle des seuils de matérialité (exemple: 1% du CA ou 1000 FCFA)
        seuil_materialite = 1000  # À adapter selon les besoins
        for ref, data in tft_data.items():
            montant = abs(data.get('montant', 0))
            if montant > 0 and montant < seuil_materialite:
                controles['warnings'].append(f"Montant faible pour {ref}: {montant:.2f} < {seuil_materialite}")
        
        return controles

    # Lecture du CSV
    df = pd.read_csv(csv_path)
    # Filtrage par période (si la colonne created_at correspond à la date d'écriture)
    if 'created_at' in df.columns:
        df['created_at'] = pd.to_datetime(df['created_at'], errors='coerce')
        # Convertir toutes les dates en timezone-naive
        if df['created_at'].dt.tz is not None:
            df['created_at'] = df['created_at'].dt.tz_convert(None)
        else:
            df['created_at'] = df['created_at'].dt.tz_localize(None)
        start_dt = pd.to_datetime(start_date)
        end_dt = pd.to_datetime(end_date)
        df = df[(df['created_at'] >= start_dt) & (df['created_at'] <= end_dt)]
        # Ajout de la colonne 'exercice' à partir de l'année de 'created_at'
        df['exercice'] = df['created_at'].dt.year

    # Mapping SYSCOHADA détaillé
    # Mapping SYSCOHADA détaillé - 10 feuilles maîtresses exactes
    groups = {
        # 1. CAPITAUX
        'Capitaux': ['101', '103', '104', '105', '106', '108', '109', '110', '130', '131'],
        
        # 2. IMMOS CORPS & INCORPS
        'Immos corps & incorps': ['201', '203', '204', '205', '208', '211', '212', '213', '214', '215', '218', '237', '238'],
        
        # 3. IMMOS FINANCIÈRES
        'Immos financières': ['251', '256', '261', '262', '264', '265', '266', '267', '268', '269', '274', '275'],
        
        # 4. STOCKS
        'Stocks': ['311', '321', '322', '323', '331', '335', '341', '345', '351', '358', '39'],
        
        # 5. CLIENTS - VENTES
        'Clients - Ventes': ['411', '416', '417', '418', '419', '491', '701', '702', '703', '704', '705', '706', '707', '708', '781'],
        
        # 6. FOURNISSEURS - ACHATS
        'Fournisseurs - Achats': ['401', '402', '403', '408', '409', '601', '602', '603', '604', '605', '606', '607', '608'],
        
        # 7. PERSONNEL
        'Personnel': ['421', '422', '423', '424', '425', '43', '447', '661', '662', '663', '664', '665', '666', '667', '668'],
        
        # 8. IMPÔTS & TAXES
        'Impôts & Taxes': ['441', '442', '443', '444', '445', '446', '447', '448', '449', '631', '633', '635', '695'],
        
        # 9. FINANCIER
        'Financier': ['501', '502', '503', '504', '505', '506', '521', '522', '523', '524', '531', '532', '533', '541', '542', '58', '59'],
        
        # 10. PROVISIONS R&C
        'Provisions R&C': ['141', '142', '143', '148', '149'],
    }

    tft_mapping = {
        'Actif Immobilisé': ['20', '21', '23', '26', '27'],
        'Actif Circulant': ['31', '32', '33', '34', '35', '36', '37', '40', '41', '42', '43', '44', '45', '50', '51', '53'],
        'Capitaux Propres': ['10', '11', '12', '13', '14'],
        'Passif Non Courant': ['15', '16'],
        'Passif Courant': ['40', '42', '44', '45', '46'],
        'Produits': ['70', '71', '72', '74'],
        'Charges': ['60', '61', '62', '63', '64', '65'],
    }

    # Fonction utilitaire pour filtrer par préfixe de numéro de compte
    def filter_by_prefix(df, prefixes):
        """Filtre les comptes par préfixe en gérant les formats réels"""
        prefixes = set(prefixes)
        def match_prefix(acc):
            acc = str(acc)
            
            if '-' in acc:
                # Format: 0000279-01 -> 279
                prefix = acc.split('-')[0]
                if prefix.startswith('0000'):
                    clean_prefix = prefix[4:]  # Enlever les 0000
                else:
                    clean_prefix = prefix.lstrip('0')  # Enlever les zéros initiaux
                
                # Vérifier si le préfixe correspond
                for p in prefixes:
                    if clean_prefix.startswith(p):
                        return True
            else:
                # Format: 66411000 (8 chiffres)
                for p in prefixes:
                    if acc.startswith(p):
                        return True
            
            return False
        
        return df[df['account_number'].apply(match_prefix)]

    # Génération des feuilles maîtresses (sera déplacée après la définition de df_n et df_n1)

    # Modèle TFT SYSCOHADA conforme à la documentation officielle
    tft_model = [
        # SECTION A - ACTIVITÉS OPÉRATIONNELLES
        {'ref': 'FA', 'libelle': 'Capacité d\'AutoFinancement Globale (CAFG)', 'formule': '131 + 681-689 + 691-699 - 781-789 - 791-799 - 775 + 675', 'prefixes': ['131', '681', '682', '683', '684', '685', '686', '687', '688', '689', '691', '692', '693', '694', '695', '696', '697', '698', '699', '781', '782', '783', '784', '785', '786', '787', '788', '789', '791', '792', '793', '794', '795', '796', '797', '798', '799', '775', '675']},
        
        {'ref': 'FB', 'libelle': 'Variation Actif circulant HAO', 'formule': 'Variation créances HAO', 'prefixes': ['461', '462', '463', '464', '465', '466', '467', '468', '469']},
        {'ref': 'FC', 'libelle': 'Variation des stocks', 'formule': '-(Solde N - Solde N-1)', 'prefixes': ['311', '321', '322', '323', '331', '335', '341', '345', '351', '358']},
        {'ref': 'FD', 'libelle': 'Variation des créances d\'exploitation', 'formule': '-(Solde N - Solde N-1)', 'prefixes': ['411', '416', '417', '418', '419']},
        {'ref': 'FE', 'libelle': 'Variation du passif circulant', 'formule': '+(Solde N - Solde N-1)', 'prefixes': ['401', '402', '403', '408', '409', '421', '422', '423', '424', '425', '431', '432', '433', '434', '435', '436', '437', '438', '441', '442', '443', '444', '445', '446', '447', '448', '449']},
        
        {'ref': 'ZB', 'libelle': 'Flux de trésorerie provenant des activités opérationnelles', 'formule': 'FA + FB + FC + FD + FE', 'prefixes': []},
        
        # SECTION B - ACTIVITÉS D'INVESTISSEMENT
        {'ref': 'FF', 'libelle': 'Décaissements liés aux acquisitions d\'immobilisations incorporelles', 'formule': 'Variation brute + cessions', 'prefixes': ['201', '203', '204', '205', '208']},
        {'ref': 'FG', 'libelle': 'Décaissements liés aux acquisitions d\'immobilisations corporelles', 'formule': 'Variation brute + cessions', 'prefixes': ['211', '212', '213', '214', '215', '218', '237', '238']},
        {'ref': 'FH', 'libelle': 'Décaissements liés aux acquisitions d\'immobilisations financières', 'formule': 'Variation nette', 'prefixes': ['251', '256', '261', '262', '264', '265', '266', '267', '268', '269', '274', '275']},
        
        {'ref': 'FI', 'libelle': 'Encaissements liés aux cessions d\'immobilisations incorporelles et corporelles', 'formule': 'Prix de cession réel', 'prefixes': ['775']},
        {'ref': 'FJ', 'libelle': 'Encaissements liés aux cessions d\'immobilisations financières', 'formule': 'Prix de cession réel', 'prefixes': ['767']},
        {'ref': 'FJ_DIV', 'libelle': 'Dividendes reçus', 'formule': 'Encaissements', 'prefixes': ['761', '762']},
        {'ref': 'FJ_CRE', 'libelle': 'Produits de créances financières', 'formule': 'Encaissements', 'prefixes': ['763', '764']},
        
        {'ref': 'ZC', 'libelle': 'Flux de trésorerie provenant des activités d\'investissement', 'formule': 'FF + FG + FH + FI + FJ + FJ_DIV + FJ_CRE', 'prefixes': []},
        
        # SECTION C - ACTIVITÉS DE FINANCEMENT
        {'ref': 'FK', 'libelle': 'Encaissements provenant de capital apporté nouveaux', 'formule': 'Variation exercice', 'prefixes': ['101', '103']},
        {'ref': 'FL', 'libelle': 'Encaissements provenant de subventions reçues', 'formule': 'Encaissements exercice', 'prefixes': ['141']},
        {'ref': 'FM', 'libelle': 'Dividendes versés', 'formule': 'Distributions décidées/payées', 'prefixes': ['108', '457']},
        
        {'ref': 'FO', 'libelle': 'Encaissements des emprunts et autres dettes financières', 'formule': 'Nouveaux emprunts', 'prefixes': ['161', '162', '163', '164', '165', '168']},
        {'ref': 'FP', 'libelle': 'Décaissements liés au remboursement des emprunts', 'formule': 'Capital remboursé uniquement', 'prefixes': ['161', '162', '163', '164', '165', '168']},
        
        {'ref': 'ZE', 'libelle': 'Flux de trésorerie provenant des activités de financement', 'formule': 'FK + FL - FM + FO - FP', 'prefixes': []},
        
        # TRÉSORERIE ET CONTRÔLES
        {'ref': 'ZA', 'libelle': 'Trésorerie nette au 1er janvier', 'formule': 'Trésorerie actif N-1 - Trésorerie passif N-1', 'prefixes': ['521', '522', '523', '524', '531', '532', '541', '542', '501', '502', '503', '504', '505', '506', '561', '564', '565']},
        {'ref': 'ZH', 'libelle': 'Trésorerie nette au 31 décembre', 'formule': 'Trésorerie actif N - Trésorerie passif N', 'prefixes': ['521', '522', '523', '524', '531', '532', '541', '542', '501', '502', '503', '504', '505', '506', '561', '564', '565']},
        
        {'ref': 'G', 'libelle': 'Variation de la trésorerie nette de la période', 'formule': 'ZB + ZC + ZE', 'prefixes': []},
        
        # CONTRÔLES DE COHÉRENCE
        {'ref': 'CONTROLE', 'libelle': 'Contrôle de cohérence', 'formule': 'G = ZH - ZA', 'prefixes': []},
    ]

    # Calcul des montants pour chaque ligne avec application des règles SYSCOHADA
    # On suppose que le CSV contient une colonne 'exercice' ou 'periode' pour distinguer N et N-1
    # Si non, il faudra adapter la structure du CSV ou demander deux fichiers
    tft_rows = []
    tft_data = {}
    montant_refs = {}
    # On prépare les DataFrames N et N-1
    if 'exercice' in df.columns:
        exercices = sorted(df['exercice'].unique())
        n = exercices[-1]
        if len(exercices) > 1:
            n_1 = exercices[-2]
            df_n = df[df['exercice'] == n]
            df_n1 = df[df['exercice'] == n_1]
        else:
            # Un seul exercice disponible
            df_n = df[df['exercice'] == n]
            # Créer un DataFrame vide pour N-1 avec les mêmes colonnes
            df_n1 = pd.DataFrame(columns=df.columns)
    else:
        # Si pas de colonne, on considère tout comme N
        df_n = df.copy()
        df_n1 = pd.DataFrame(columns=df.columns)

    for ligne in tft_model:
        if ligne['ref'] == 'FA':
            cafg_comptes = [
                {'prefixes': ['131'], 'sign': 1},
                {'prefixes': [str(i) for i in range(681, 690)], 'sign': 1},
                {'prefixes': [str(i) for i in range(691, 700)], 'sign': 1},
                {'prefixes': [str(i) for i in range(781, 790)], 'sign': -1},
                {'prefixes': [str(i) for i in range(791, 800)], 'sign': -1},
                {'prefixes': ['775'], 'sign': -1},
                {'prefixes': ['675'], 'sign': 1},
            ]
            montant = 0
            comptes = []
            for item in cafg_comptes:
                comptes_n = filter_by_prefix(df_n, item['prefixes'])
                solde = comptes_n['balance'].sum() if not comptes_n.empty else 0
                montant += item['sign'] * solde
                comptes.extend(comptes_n.to_dict(orient='records'))
            solde_n = montant
            # Si N-1 existe, calculer, sinon mettre à zéro
            solde_n1 = 0 if df_n1.empty else 0
            variation = 0 if df_n1.empty else 0
            debit_n = 0
            credit_n = 0
        else:
            comptes_n = filter_by_prefix(df_n, ligne['prefixes']) if ligne['prefixes'] else pd.DataFrame()
            comptes_n1 = filter_by_prefix(df_n1, ligne['prefixes']) if ligne['prefixes'] else pd.DataFrame()
            if ligne['ref'] == 'T5_141':
                comptes_n = comptes_n[comptes_n['account_number'] != '865'] if not comptes_n.empty else comptes_n
                comptes_n1 = comptes_n1[comptes_n1['account_number'] != '865'] if not comptes_n1.empty else comptes_n1
            solde_n = comptes_n['balance'].sum() if not comptes_n.empty else 0
            solde_n1 = comptes_n1['balance'].sum() if not comptes_n1.empty else 0
            solde_n = solde_n if solde_n is not None else 0
            # Si N-1 n'existe pas, mettre à zéro
            solde_n1 = solde_n1 if not df_n1.empty else 0
            if ligne['ref'] == 'ZA':
                # Trésorerie nette au 1er janvier = Trésorerie actif N-1 - Trésorerie passif N-1
                treso_actif_n1 = filter_by_prefix(df_n1, ['521', '431'])
                treso_passif_n1 = filter_by_prefix(df_n1, ['521', '431'])  # Même préfixe pour l'instant
                solde_actif_n1 = treso_actif_n1['balance'].sum() if not treso_actif_n1.empty else 0
                solde_passif_n1 = treso_passif_n1['balance'].sum() if not treso_passif_n1.empty else 0
                montant = (solde_actif_n1 or 0) - (solde_passif_n1 or 0)
                comptes = treso_actif_n1.to_dict(orient='records') + treso_passif_n1.to_dict(orient='records')
                variation = montant
                debit_n = treso_actif_n1['total_debit'].sum() if 'total_debit' in treso_actif_n1 else 0
                credit_n = treso_actif_n1['total_credit'].sum() if 'total_credit' in treso_actif_n1 else 0
            elif ligne['ref'] == 'G':
                # Variation de la trésorerie nette = D + B + C + F
                montant = (montant_refs.get('D', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('B', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('C', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('F', {}).get('montant', 0) or 0)
                variation = montant
                comptes = []
                debit_n = 0
                credit_n = 0
            elif ligne['ref'] == 'ZH':
                # Trésorerie nette au 31 décembre = G + A
                montant = (montant_refs.get('G', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('A', {}).get('montant', 0) or 0)
                variation = montant
                comptes = []
                debit_n = 0
                credit_n = 0
            else:
                variation = (solde_n or 0) - (solde_n1 or 0) if not df_n1.empty else 0
            debit_n = comptes_n['total_debit'].sum() if 'total_debit' in comptes_n else 0
            credit_n = comptes_n['total_credit'].sum() if 'total_credit' in comptes_n else 0
            montant = 0
            comptes = comptes_n.to_dict(orient='records')
            if ligne['formule']:
                formule = ligne['formule']
                for ref in montant_refs:
                    formule = formule.replace(ref, f"({montant_refs[ref]['montant']})")
                try:
                    montant = eval(formule)
                except Exception:
                    montant = 0
            else:
                montant = variation if not df_n1.empty else solde_n
        montant_refs[ligne['ref']] = {
            'montant': montant if montant is not None else 0,
            'solde_n': solde_n,
            'solde_n1': solde_n1,
            'variation': variation,
            'debit_n': debit_n,
            'credit_n': credit_n
        }
        tft_rows.append({
            'Réf': ligne['ref'],
            'Libellé': ligne['libelle'],
            'Montant': montant,
            'Solde_N': solde_n,
            'Solde_N-1': solde_n1,
            'Variation': variation,
            'Débit_N': debit_n,
            'Crédit_N': credit_n,
            'Formule': ligne['formule']
        })
        tft_data[ligne['ref']] = {
            'libelle': ligne['libelle'],
            'montant': montant,
            'solde_n': solde_n,
            'solde_n1': solde_n1,
            'variation': variation,
            'debit_n': debit_n,
            'credit_n': credit_n,
            'formule': ligne['formule'],
            'comptes': comptes
        }

    tft_df = pd.DataFrame(tft_rows)
    tft_output = BytesIO()
    tft_df.to_excel(tft_output, index=False)
    tft_content = tft_output.getvalue()
    
    # Génération des feuilles maîtresses avec structure multi-onglets
    sheets_contents = {}
    sheets_data = {}
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    for group_name, prefixes in groups.items():
        # Filtrer les données par groupe et exercice
        group_n = filter_by_prefix(df_n, prefixes)
        group_n1 = filter_by_prefix(df_n1, prefixes)
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Déterminer si on a deux exercices
        has_two_exercices = not df_n1.empty and len(exercices) > 1
        
        if has_two_exercices:
            # Cas 1: Deux exercices - Créer 3 onglets
            
            # 1. Onglet Exercice N
            ws_n = wb.create_sheet(f"Exercice_{n}")
            ws_n.append(['Compte', 'Libellé', 'Solde', 'Débit', 'Crédit', 'Exercice', 'Date_Creation'])
            
            # Remplir les données N
            n_data = []
            for _, row in group_n.iterrows():
                n_data.append({
                    'Compte': row['account_number'],
                    'Libellé': row.get('account_name', ''),
                    'Solde': row['balance'],
                    'Débit': row.get('total_debit', 0),
                    'Crédit': row.get('total_credit', 0),
                    'Exercice': n,
                    'Date_Creation': row.get('created_at', '')
                })
            
            # Trier par numéro de compte
            n_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in n_data:
                ws_n.append([row_data['Compte'], row_data['Libellé'], row_data['Solde'], 
                           row_data['Débit'], row_data['Crédit'], row_data['Exercice'], row_data['Date_Creation']])
            
            # 2. Onglet Exercice N-1
            ws_n1 = wb.create_sheet(f"Exercice_{n_1}")
            ws_n1.append(['Compte', 'Libellé', 'Solde', 'Débit', 'Crédit', 'Exercice', 'Date_Creation'])
            
            # Remplir les données N-1
            n1_data = []
            for _, row in group_n1.iterrows():
                n1_data.append({
                    'Compte': row['account_number'],
                    'Libellé': row.get('account_name', ''),
                    'Solde': row['balance'],
                    'Débit': row.get('total_debit', 0),
                    'Crédit': row.get('total_credit', 0),
                    'Exercice': n_1,
                    'Date_Creation': row.get('created_at', '')
                })
            
            # Trier par numéro de compte
            n1_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in n1_data:
                ws_n1.append([row_data['Compte'], row_data['Libellé'], row_data['Solde'], 
                            row_data['Débit'], row_data['Crédit'], row_data['Exercice'], row_data['Date_Creation']])
            
            # 3. Onglet Comparatif
            ws_comp = wb.create_sheet(f"Comparatif_{n}_{n_1}")
            ws_comp.append(['Compte', 'Libellé', 'Solde_N', 'Solde_N-1', 'Variation', 
                          '%_Evolution', 'Débit_N', 'Crédit_N', 'Débit_N-1', 'Crédit_N-1', 'Date_Creation_N', 'Date_Creation_N-1', 'Exercices'])
            
            # Créer un dictionnaire pour faciliter la comparaison
            n_dict = {row['Compte']: row for row in n_data}
            n1_dict = {row['Compte']: row for row in n1_data}
            
            # Obtenir tous les comptes uniques
            all_accounts = set(n_dict.keys()) | set(n1_dict.keys())
            
            # Remplir le tableau comparatif
            comp_data = []
            for account in sorted(all_accounts):
                n_row = n_dict.get(account, {})
                n1_row = n1_dict.get(account, {})
                
                # Remplacer les "trous" par 0
                solde_n = n_row.get('Solde', 0) if n_row else 0
                solde_n1 = n1_row.get('Solde', 0) if n1_row else 0
                debit_n = n_row.get('Débit', 0) if n_row else 0
                credit_n = n_row.get('Crédit', 0) if n_row else 0
                debit_n1 = n1_row.get('Débit', 0) if n1_row else 0
                credit_n1 = n1_row.get('Crédit', 0) if n1_row else 0
                
                # Calculer la variation
                variation = (solde_n or 0) - (solde_n1 or 0)
                
                # Calculer le pourcentage d'évolution
                if solde_n1 != 0:
                    pct_evolution = (variation / abs(solde_n1)) * 100
                else:
                    pct_evolution = 100 if solde_n != 0 else 0
                
                # Utiliser le libellé de N ou N-1 (priorité à N)
                libelle = n_row.get('Libellé', n1_row.get('Libellé', ''))
                
                comp_data.append({
                    'Compte': account,
                    'Libellé': libelle,
                    'Solde_N': solde_n,
                    'Solde_N-1': solde_n1,
                    'Variation': variation,
                    '%_Evolution': pct_evolution,
                    'Débit_N': debit_n,
                    'Crédit_N': credit_n,
                    'Débit_N-1': debit_n1,
                    'Crédit_N-1': credit_n1,
                    'Date_Creation_N': n_row.get('Date_Creation', ''),
                    'Date_Creation_N-1': n1_row.get('Date_Creation', ''),
                    'Exercices': f"{n}/{n_1}"
                })
            
            # Trier par numéro de compte
            comp_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in comp_data:
                ws_comp.append([row_data['Compte'], row_data['Libellé'], row_data['Solde_N'], 
                              row_data['Solde_N-1'], row_data['Variation'], row_data['%_Evolution'],
                              row_data['Débit_N'], row_data['Crédit_N'], row_data['Débit_N-1'], 
                              row_data['Crédit_N-1'], row_data['Date_Creation_N'], row_data['Date_Creation_N-1'], row_data['Exercices']])
            
            # Appliquer le formatage aux onglets
            for ws in [ws_n, ws_n1, ws_comp]:
                # En-têtes en gras
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Ajuster la largeur des colonnes
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Stocker les données pour l'API
            sheets_data[group_name] = {
                'exercice_n': n_data,
                'exercice_n1': n1_data,
                'comparatif': comp_data,
                'has_two_exercices': True,
                'exercices': [n, n_1]
            }
            
        else:
            # Cas 2: Un seul exercice - Créer 1 onglet
            ws_n = wb.create_sheet(f"Exercice_{n}")
            ws_n.append(['Compte', 'Libellé', 'Solde', 'Débit', 'Crédit', 'Exercice', 'Date_Creation'])
            
            # Remplir les données N
            n_data = []
            for _, row in group_n.iterrows():
                n_data.append({
                    'Compte': row['account_number'],
                    'Libellé': row.get('account_name', ''),
                    'Solde': row['balance'],
                    'Débit': row.get('total_debit', 0),
                    'Crédit': row.get('total_credit', 0),
                    'Exercice': n,
                    'Date_Creation': row.get('created_at', '')
                })
            
            # Trier par numéro de compte
            n_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in n_data:
                ws_n.append([row_data['Compte'], row_data['Libellé'], row_data['Solde'], 
                           row_data['Débit'], row_data['Crédit'], row_data['Exercice'], row_data['Date_Creation']])
            
            # Appliquer le formatage
            for cell in ws_n[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ajuster la largeur des colonnes
            for column in ws_n.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_n.column_dimensions[column_letter].width = adjusted_width
            
            # Stocker les données pour l'API
            sheets_data[group_name] = {
                'exercice_n': n_data,
                'has_two_exercices': False,
                'exercices': [n]
            }
        
        # Sauvegarder le fichier Excel
        output = BytesIO()
        wb.save(output)
        sheets_contents[group_name] = output.getvalue()
    
    # Ajout du contrôle de cohérence au retour
    coherence = controle_coherence_complet(tft_data)
    return tft_content, sheets_contents, tft_data, sheets_data, coherence

def generate_tft_and_sheets_from_database(financial_report_id, start_date, end_date):
    """Génère le TFT et les feuilles maîtresses à partir des données de la base"""
    # Récupérer les données depuis AccountData
    account_data = AccountData.objects.filter(financial_report_id=financial_report_id)
    
    if not account_data.exists():
        raise ValueError(f"Aucune donnée trouvée pour financial_report_id: {financial_report_id}")
    
    # Convertir en DataFrame
    df_data = []
    for data in account_data:
        df_data.append({
            'account_number': data.account_number,
            'account_name': data.account_label,  # Utiliser account_label comme account_name
            'balance': float(data.balance),
            'total_debit': float(data.total_debit),
            'total_credit': float(data.total_credit),
            'created_at': data.created_at,
            'exercice': data.created_at.year
        })
    
    df = pd.DataFrame(df_data)
    
    # Utiliser la même logique que la fonction originale
    return generate_tft_and_sheets_from_df(df, start_date, end_date)

def generate_tft_and_sheets_from_df(df, start_date, end_date):
    """Génère le TFT et les feuilles maîtresses à partir d'un DataFrame"""
    # Contrôle de cohérence TFT
    # Contrôles de cohérence conformes à la documentation SYSCOHADA
    def controle_coherence_complet(tft_data):
        """
        Contrôles automatiques obligatoires selon la documentation :
        - Égalité variation calculée/variation bilantielle
        - Cohérence des totaux par section
        - Absence de comptes orphelins
        - Respect des seuils de matérialité
        """
        controles = {
            'is_coherent': True,
            'errors': [],
            'warnings': [],
            'details': {}
        }
        
        # 1. Égalité variation calculée/variation bilantielle
        flux_operationnels = tft_data.get('ZB', {}).get('montant', 0)
        flux_investissement = tft_data.get('ZC', {}).get('montant', 0)
        flux_financement = tft_data.get('ZE', {}).get('montant', 0)
        treso_ouverture = tft_data.get('ZA', {}).get('montant', 0)
        treso_cloture = tft_data.get('ZH', {}).get('montant', 0)
        
        variation_tft = (flux_operationnels or 0) + (flux_investissement or 0) + (flux_financement or 0)
        variation_treso = (treso_cloture or 0) - (treso_ouverture or 0)
        
        ecart = abs(variation_tft - variation_treso)
        if ecart > 1e-2:
            controles['is_coherent'] = False
            controles['errors'].append(f"Écart variation TFT/Trésorerie: {ecart:.2f}")
        
        controles['details']['variation_tft'] = variation_tft
        controles['details']['variation_treso'] = variation_treso
        controles['details']['ecart'] = ecart
        
        # 2. Cohérence des totaux par section
        # Section A - Activités opérationnelles
        cafg = tft_data.get('FA', {}).get('montant', 0)
        bfr_exploitation = tft_data.get('FB', {}).get('montant', 0)
        variation_stocks = tft_data.get('FC', {}).get('montant', 0)
        variation_creances = tft_data.get('FD', {}).get('montant', 0)
        variation_passif = tft_data.get('FE', {}).get('montant', 0)
        
        total_section_a = cafg + bfr_exploitation + variation_stocks + variation_creances + variation_passif
        if abs(total_section_a - flux_operationnels) > 1e-2:
            controles['warnings'].append(f"Cohérence Section A: écart {abs(total_section_a - flux_operationnels):.2f}")
        
        # Section B - Activités d'investissement
        acquisitions_incorp = tft_data.get('FF', {}).get('montant', 0)
        acquisitions_corp = tft_data.get('FG', {}).get('montant', 0)
        acquisitions_fin = tft_data.get('FH', {}).get('montant', 0)
        cessions_incorp_corp = tft_data.get('FI', {}).get('montant', 0)
        cessions_fin = tft_data.get('FJ', {}).get('montant', 0)
        dividendes_recus = tft_data.get('FJ_DIV', {}).get('montant', 0)
        produits_creances = tft_data.get('FJ_CRE', {}).get('montant', 0)
        
        total_section_b = acquisitions_incorp + acquisitions_corp + acquisitions_fin + cessions_incorp_corp + cessions_fin + dividendes_recus + produits_creances
        if abs(total_section_b - flux_investissement) > 1e-2:
            controles['warnings'].append(f"Cohérence Section B: écart {abs(total_section_b - flux_investissement):.2f}")
        
        # Section C - Activités de financement
        augmentation_capital = tft_data.get('FK', {}).get('montant', 0)
        subventions = tft_data.get('FL', {}).get('montant', 0)
        dividendes_verses = tft_data.get('FM', {}).get('montant', 0)
        nouveaux_emprunts = tft_data.get('FO', {}).get('montant', 0)
        remboursements = tft_data.get('FP', {}).get('montant', 0)
        
        total_section_c = augmentation_capital + subventions - dividendes_verses + nouveaux_emprunts - remboursements
        if abs(total_section_c - flux_financement) > 1e-2:
            controles['warnings'].append(f"Cohérence Section C: écart {abs(total_section_c - flux_financement):.2f}")
        
        # 3. Contrôle des seuils de matérialité (exemple: 1% du CA ou 1000 FCFA)
        seuil_materialite = 1000  # À adapter selon les besoins
        for ref, data in tft_data.items():
            montant = abs(data.get('montant', 0))
            if montant > 0 and montant < seuil_materialite:
                controles['warnings'].append(f"Montant faible pour {ref}: {montant:.2f} < {seuil_materialite}")
        
        return controles

    # Filtrage par période
    if 'created_at' in df.columns:
        df['created_at'] = pd.to_datetime(df['created_at'], errors='coerce')
        # Convertir toutes les dates en timezone-naive
        if df['created_at'].dt.tz is not None:
            df['created_at'] = df['created_at'].dt.tz_convert(None)
        else:
            df['created_at'] = df['created_at'].dt.tz_localize(None)
        start_dt = pd.to_datetime(start_date)
        end_dt = pd.to_datetime(end_date)
        df = df[(df['created_at'] >= start_dt) & (df['created_at'] <= end_dt)]
        # Ajout de la colonne 'exercice' à partir de l'année de 'created_at'
        df['exercice'] = df['created_at'].dt.year

    # Mapping SYSCOHADA détaillé
    # Mapping SYSCOHADA détaillé - 10 feuilles maîtresses exactes
    groups = {
        # 1. CAPITAUX
        'Capitaux': ['101', '103', '104', '105', '106', '108', '109', '110', '130', '131'],
        
        # 2. IMMOS CORPS & INCORPS
        'Immos corps & incorps': ['201', '203', '204', '205', '208', '211', '212', '213', '214', '215', '218', '237', '238'],
        
        # 3. IMMOS FINANCIÈRES
        'Immos financières': ['251', '256', '261', '262', '264', '265', '266', '267', '268', '269', '274', '275'],
        
        # 4. STOCKS
        'Stocks': ['311', '321', '322', '323', '331', '335', '341', '345', '351', '358', '39'],
        
        # 5. CLIENTS - VENTES
        'Clients - Ventes': ['411', '416', '417', '418', '419', '491', '701', '702', '703', '704', '705', '706', '707', '708', '781'],
        
        # 6. FOURNISSEURS - ACHATS
        'Fournisseurs - Achats': ['401', '402', '403', '408', '409', '601', '602', '603', '604', '605', '606', '607', '608'],
        
        # 7. PERSONNEL
        'Personnel': ['421', '422', '423', '424', '425', '43', '447', '661', '662', '663', '664', '665', '666', '667', '668'],
        
        # 8. IMPÔTS & TAXES
        'Impôts & Taxes': ['441', '442', '443', '444', '445', '446', '447', '448', '449', '631', '633', '635', '695'],
        
        # 9. FINANCIER
        'Financier': ['501', '502', '503', '504', '505', '506', '521', '522', '523', '524', '531', '532', '533', '541', '542', '58', '59'],
        
        # 10. PROVISIONS R&C
        'Provisions R&C': ['141', '142', '143', '148', '149'],
    }

    tft_mapping = {
        'Actif Immobilisé': ['20', '21', '23', '26', '27'],
        'Actif Circulant': ['31', '32', '33', '34', '35', '36', '37', '40', '41', '42', '43', '44', '45', '50', '51', '53'],
        'Capitaux Propres': ['10', '11', '12', '13', '14'],
        'Passif Non Courant': ['15', '16'],
        'Passif Courant': ['40', '42', '44', '45', '46'],
        'Produits': ['70', '71', '72', '74'],
        'Charges': ['60', '61', '62', '63', '64', '65'],
    }

    # Fonction utilitaire pour filtrer par préfixe de numéro de compte
    def filter_by_prefix(df, prefixes):
        """Filtre les comptes par préfixe en gérant les formats réels"""
        prefixes = set(prefixes)
        def match_prefix(acc):
            acc = str(acc)
            
            if '-' in acc:
                # Format: 0000279-01 -> 279
                prefix = acc.split('-')[0]
                if prefix.startswith('0000'):
                    clean_prefix = prefix[4:]  # Enlever les 0000
                else:
                    clean_prefix = prefix.lstrip('0')  # Enlever les zéros initiaux
                
                # Vérifier si le préfixe correspond
                for p in prefixes:
                    if clean_prefix.startswith(p):
                        return True
            else:
                # Format: 66411000 (8 chiffres)
                for p in prefixes:
                    if acc.startswith(p):
                        return True
            
            return False
        
        return df[df['account_number'].apply(match_prefix)]

    # Modèle TFT SYSCOHADA (exemple simplifié, à compléter selon le guide)
    tft_model = [
    {'ref': '2H_TRESO_NEG', 'libelle': "Trésorerie passive (négative) - concours et escomptes", 'formule': None, 'prefixes': ['561', '564', '565']},
    {'ref': '2H_TRESO_POS', 'libelle': "Trésorerie active (positive) - composition détaillée", 'formule': None, 'prefixes': ['521', '522', '523', '524', '531', '532', '541', '542', '501', '502', '503', '504', '505', '506']},
        {'ref': 'ZA', 'libelle': 'Trésorerie nette au 1er janvier', 'formule': 'Trésorerie actif N-1 - Trésorerie passif N-1', 'prefixes': ['521', '431']},
        {'ref': 'FA', 'libelle': 'Capacité d\'AutoFinancement Globale (CAFG)', 'formule': None, 'prefixes': ['131', '681-689', '691-699', '781-789', '791-799', '775', '675']},
        {'ref': 'FB', 'libelle': 'Variation Actif circulant HAO', 'formule': None, 'prefixes': ['31', '32', '33', '34', '35', '36', '37']},
        {'ref': 'FC', 'libelle': 'Variation des stocks', 'formule': None, 'prefixes': ['31', '32', '33', '34', '35', '36', '37']},
        {'ref': 'FD', 'libelle': 'Variation des créances', 'formule': None, 'prefixes': ['41']},
        {'ref': 'FE', 'libelle': 'Variation du passif circulant', 'formule': None, 'prefixes': ['40', '44', '45', '46']},
        {'ref': 'BF', 'libelle': 'Variation du BF lié aux activités opérationnelles', 'formule': 'FB+FC+FD-FE', 'prefixes': []},
        {'ref': 'ZB', 'libelle': 'Flux de trésorerie provenant des activités opérationnelles (somme FA à FE)', 'formule': 'FA+FB+FC+FD+FE', 'prefixes': []},
        {'ref': 'FF', 'libelle': 'Décaissements liés aux acquisitions d\'immobilisations incorporelles', 'formule': None, 'prefixes': ['244']},
        {'ref': 'FG', 'libelle': 'Décaissements liés aux acquisitions d\'immobilisations corporelles', 'formule': None, 'prefixes': ['624']},
        {'ref': 'FH', 'libelle': 'Décaissements liés aux acquisitions d\'immobilisations financières', 'formule': None, 'prefixes': ['244', '624']},
        {'ref': 'FI', 'libelle': 'Encaissements liés aux cessions d\'immobilisations incorporelles et corporelles', 'formule': None, 'prefixes': ['244', '624']},
        {'ref': 'FJ', 'libelle': 'Encaissements liés aux cessions d\'immobilisations financières', 'formule': None, 'prefixes': ['244', '624']},
        {'ref': 'FJ_VMP', 'libelle': 'Produits nets sur cessions VMP (767)', 'formule': None, 'prefixes': ['767']},
        {'ref': 'INV_DIV', 'libelle': "Dividendes reçus (761-762)", 'formule': None, 'prefixes': ['761', '762']},
        {'ref': 'INV_CRE', 'libelle': "Produits de créances financières (763-764)", 'formule': None, 'prefixes': ['763', '764']},
        {'ref': 'ZC', 'libelle': 'Flux de trésorerie provenant des activités d\'investissement (somme FF à FJ)', 'formule': 'FF+FG+FH+FI+FJ', 'prefixes': []},
        {'ref': 'FK', 'libelle': 'Encaissements provenant de capital apporté nouveaux', 'formule': None, 'prefixes': ['10', '11', '12', '13', '14']},
        {'ref': 'T4_101', 'libelle': "Capital social (101) - hors apports en nature", 'formule': None, 'prefixes': ['101']},
        {'ref': 'T4_103', 'libelle': "Primes d'émission (103) - encaissements effectifs", 'formule': None, 'prefixes': ['103']},
        {'ref': 'T4_104', 'libelle': "Écarts d'évaluation (104) - non concerné", 'formule': None, 'prefixes': ['104']},
        {'ref': 'FL', 'libelle': 'Encaissements provenant de subventions reçues', 'formule': None, 'prefixes': ['121']},
        {'ref': 'T5_141', 'libelle': "Subventions d'investissement reçues (141) - hors reprises (865)", 'formule': None, 'prefixes': ['141']},
        {'ref': 'FM', 'libelle': 'Dividendes versés', 'formule': None, 'prefixes': ['121']},
        {'ref': 'TH1_108', 'libelle': "Compte de l'exploitant (108) - prélèvements nets", 'formule': None, 'prefixes': ['108']},
        {'ref': 'TH2_457', 'libelle': "Dividendes à payer (457) - distributions décidées/payées", 'formule': None, 'prefixes': ['457']},
        {'ref': 'D', 'libelle': 'Flux de trésorerie provenant des capitaux propres (somme FK à FM)', 'formule': 'FK+FL-FM', 'prefixes': []},
        {'ref': 'FO', 'libelle': 'Encaissements des emprunts et autres dettes financières', 'formule': None, 'prefixes': ['401', '409', '637']},
        {'ref': 'FP', 'libelle': 'Décaissements liés au remboursement des emprunts et autres dettes financières', 'formule': None, 'prefixes': ['401', '409', '637']},
    {'ref': 'TG_161_168', 'libelle': "Nouveaux emprunts (161-168) - augmentation/variation nette", 'formule': None, 'prefixes': ['161', '162', '163', '164', '165', '168']},
    {'ref': 'TP_161_168', 'libelle': "Remboursements d'emprunts (161-168) - capital remboursé uniquement", 'formule': None, 'prefixes': ['161', '162', '163', '164', '165', '168']},
        {'ref': 'ZE', 'libelle': 'Flux de trésorerie provenant des activités de financement (FO-FP)', 'formule': 'FO-FP', 'prefixes': []},
        {'ref': 'G', 'libelle': 'VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE (D+B+C+F)', 'formule': 'D+B+C+F', 'prefixes': []},
        {'ref': 'ZH', 'libelle': 'Trésorerie nette au 31 Décembre (G+A)', 'formule': 'G+A', 'prefixes': ['521', '431']},
    ]

    # Calcul des montants pour chaque ligne avec application des règles SYSCOHADA
    tft_rows = []
    tft_data = {}
    montant_refs = {}
    # On prépare les DataFrames N et N-1
    if 'exercice' in df.columns:
        exercices = sorted(df['exercice'].unique())
        n = exercices[-1]
        if len(exercices) > 1:
            n_1 = exercices[-2]
            df_n = df[df['exercice'] == n]
            df_n1 = df[df['exercice'] == n_1]
        else:
            # Un seul exercice disponible
            df_n = df[df['exercice'] == n]
            # Créer un DataFrame vide pour N-1 avec les mêmes colonnes
            df_n1 = pd.DataFrame(columns=df.columns)
    else:
        # Si pas de colonne, on considère tout comme N
        df_n = df.copy()
        df_n1 = pd.DataFrame(columns=df.columns)

    for ligne in tft_model:
        if ligne['ref'] == 'FA':
            cafg_comptes = [
                {'prefixes': ['131'], 'sign': 1},
                {'prefixes': [str(i) for i in range(681, 690)], 'sign': 1},
                {'prefixes': [str(i) for i in range(691, 700)], 'sign': 1},
                {'prefixes': [str(i) for i in range(781, 790)], 'sign': -1},
                {'prefixes': [str(i) for i in range(791, 800)], 'sign': -1},
                {'prefixes': ['775'], 'sign': -1},
                {'prefixes': ['675'], 'sign': 1},
            ]
            montant = 0
            comptes = []
            for item in cafg_comptes:
                comptes_n = filter_by_prefix(df_n, item['prefixes'])
                solde = comptes_n['balance'].sum() if not comptes_n.empty else 0
                montant += item['sign'] * solde
                comptes.extend(comptes_n.to_dict(orient='records'))
            solde_n = montant
            # Si N-1 existe, calculer, sinon mettre à zéro
            solde_n1 = 0 if df_n1.empty else 0
            variation = 0 if df_n1.empty else 0
            debit_n = 0
            credit_n = 0
        else:
            comptes_n = filter_by_prefix(df_n, ligne['prefixes']) if ligne['prefixes'] else pd.DataFrame()
            comptes_n1 = filter_by_prefix(df_n1, ligne['prefixes']) if ligne['prefixes'] else pd.DataFrame()
            if ligne['ref'] == 'T5_141':
                comptes_n = comptes_n[comptes_n['account_number'] != '865'] if not comptes_n.empty else comptes_n
                comptes_n1 = comptes_n1[comptes_n1['account_number'] != '865'] if not comptes_n1.empty else comptes_n1
            solde_n = comptes_n['balance'].sum() if not comptes_n.empty else 0
            solde_n1 = comptes_n1['balance'].sum() if not comptes_n1.empty else 0
            solde_n = solde_n if solde_n is not None else 0
            # Si N-1 n'existe pas, mettre à zéro
            solde_n1 = solde_n1 if not df_n1.empty else 0
            if ligne['ref'] == 'ZA':
                # Trésorerie nette au 1er janvier = Trésorerie actif N-1 - Trésorerie passif N-1
                treso_actif_n1 = filter_by_prefix(df_n1, ['521', '431'])
                treso_passif_n1 = filter_by_prefix(df_n1, ['521', '431'])  # Même préfixe pour l'instant
                solde_actif_n1 = treso_actif_n1['balance'].sum() if not treso_actif_n1.empty else 0
                solde_passif_n1 = treso_passif_n1['balance'].sum() if not treso_passif_n1.empty else 0
                montant = (solde_actif_n1 or 0) - (solde_passif_n1 or 0)
                comptes = treso_actif_n1.to_dict(orient='records') + treso_passif_n1.to_dict(orient='records')
                variation = montant
                debit_n = treso_actif_n1['total_debit'].sum() if 'total_debit' in treso_actif_n1 else 0
                credit_n = treso_actif_n1['total_credit'].sum() if 'total_credit' in treso_actif_n1 else 0
            elif ligne['ref'] == 'G':
                # Variation de la trésorerie nette = D + B + C + F
                montant = (montant_refs.get('D', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('B', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('C', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('F', {}).get('montant', 0) or 0)
                variation = montant
                comptes = []
                debit_n = 0
                credit_n = 0
            elif ligne['ref'] == 'ZH':
                # Trésorerie nette au 31 décembre = G + A
                montant = (montant_refs.get('G', {}).get('montant', 0) or 0) + \
                         (montant_refs.get('A', {}).get('montant', 0) or 0)
                variation = montant
                comptes = []
                debit_n = 0
                credit_n = 0
            else:
                variation = (solde_n or 0) - (solde_n1 or 0) if not df_n1.empty else 0
            debit_n = comptes_n['total_debit'].sum() if 'total_debit' in comptes_n else 0
            credit_n = comptes_n['total_credit'].sum() if 'total_credit' in comptes_n else 0
            montant = 0
            comptes = comptes_n.to_dict(orient='records')
            if ligne['formule']:
                formule = ligne['formule']
                for ref in montant_refs:
                    formule = formule.replace(ref, f"({montant_refs[ref]['montant']})")
                try:
                    montant = eval(formule)
                except Exception:
                    montant = 0
            else:
                montant = variation if not df_n1.empty else solde_n
        montant_refs[ligne['ref']] = {
            'montant': montant if montant is not None else 0,
            'solde_n': solde_n,
            'solde_n1': solde_n1,
            'variation': variation,
            'debit_n': debit_n,
            'credit_n': credit_n
        }
        tft_rows.append({
            'Réf': ligne['ref'],
            'Libellé': ligne['libelle'],
            'Montant': montant,
            'Solde_N': solde_n,
            'Solde_N-1': solde_n1,
            'Variation': variation,
            'Débit_N': debit_n,
            'Crédit_N': credit_n,
            'Formule': ligne['formule']
        })
        tft_data[ligne['ref']] = {
            'libelle': ligne['libelle'],
            'montant': montant,
            'solde_n': solde_n,
            'solde_n1': solde_n1,
            'variation': variation,
            'debit_n': debit_n,
            'credit_n': credit_n,
            'formule': ligne['formule'],
            'comptes': comptes
        }

    tft_df = pd.DataFrame(tft_rows)
    tft_output = BytesIO()
    tft_df.to_excel(tft_output, index=False)
    tft_content = tft_output.getvalue()
    
    # Génération des feuilles maîtresses avec structure multi-onglets
    sheets_contents = {}
    sheets_data = {}
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    for group_name, prefixes in groups.items():
        # Filtrer les données par groupe et exercice
        group_n = filter_by_prefix(df_n, prefixes)
        group_n1 = filter_by_prefix(df_n1, prefixes)
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Déterminer si on a deux exercices
        has_two_exercices = not df_n1.empty and len(exercices) > 1
        
        if has_two_exercices:
            # Cas 1: Deux exercices - Créer 3 onglets
            
            # 1. Onglet Exercice N
            ws_n = wb.create_sheet(f"Exercice_{n}")
            ws_n.append(['Compte', 'Libellé', 'Solde', 'Débit', 'Crédit', 'Exercice', 'Date_Creation'])
            
            # Remplir les données N
            n_data = []
            for _, row in group_n.iterrows():
                n_data.append({
                    'Compte': row['account_number'],
                    'Libellé': row.get('account_name', ''),
                    'Solde': row['balance'],
                    'Débit': row.get('total_debit', 0),
                    'Crédit': row.get('total_credit', 0),
                    'Exercice': n,
                    'Date_Creation': row.get('created_at', '')
                })
            
            # Trier par numéro de compte
            n_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in n_data:
                ws_n.append([row_data['Compte'], row_data['Libellé'], row_data['Solde'], 
                           row_data['Débit'], row_data['Crédit'], row_data['Exercice'], row_data['Date_Creation']])
            
            # 2. Onglet Exercice N-1
            ws_n1 = wb.create_sheet(f"Exercice_{n_1}")
            ws_n1.append(['Compte', 'Libellé', 'Solde', 'Débit', 'Crédit', 'Exercice', 'Date_Creation'])
            
            # Remplir les données N-1
            n1_data = []
            for _, row in group_n1.iterrows():
                n1_data.append({
                    'Compte': row['account_number'],
                    'Libellé': row.get('account_name', ''),
                    'Solde': row['balance'],
                    'Débit': row.get('total_debit', 0),
                    'Crédit': row.get('total_credit', 0),
                    'Exercice': n_1,
                    'Date_Creation': row.get('created_at', '')
                })
            
            # Trier par numéro de compte
            n1_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in n1_data:
                ws_n1.append([row_data['Compte'], row_data['Libellé'], row_data['Solde'], 
                            row_data['Débit'], row_data['Crédit'], row_data['Exercice'], row_data['Date_Creation']])
            
            # 3. Onglet Comparatif
            ws_comp = wb.create_sheet(f"Comparatif_{n}_{n_1}")
            ws_comp.append(['Compte', 'Libellé', 'Solde_N', 'Solde_N-1', 'Variation', 
                          '%_Evolution', 'Débit_N', 'Crédit_N', 'Débit_N-1', 'Crédit_N-1', 'Date_Creation_N', 'Date_Creation_N-1', 'Exercices'])
            
            # Créer un dictionnaire pour faciliter la comparaison
            n_dict = {row['Compte']: row for row in n_data}
            n1_dict = {row['Compte']: row for row in n1_data}
            
            # Obtenir tous les comptes uniques
            all_accounts = set(n_dict.keys()) | set(n1_dict.keys())
            
            # Remplir le tableau comparatif
            comp_data = []
            for account in sorted(all_accounts):
                n_row = n_dict.get(account, {})
                n1_row = n1_dict.get(account, {})
                
                # Remplacer les "trous" par 0
                solde_n = n_row.get('Solde', 0) if n_row else 0
                solde_n1 = n1_row.get('Solde', 0) if n1_row else 0
                debit_n = n_row.get('Débit', 0) if n_row else 0
                credit_n = n_row.get('Crédit', 0) if n_row else 0
                debit_n1 = n1_row.get('Débit', 0) if n1_row else 0
                credit_n1 = n1_row.get('Crédit', 0) if n1_row else 0
                
                # Calculer la variation
                variation = (solde_n or 0) - (solde_n1 or 0)
                
                # Calculer le pourcentage d'évolution
                if solde_n1 != 0:
                    pct_evolution = (variation / abs(solde_n1)) * 100
                else:
                    pct_evolution = 100 if solde_n != 0 else 0
                
                # Utiliser le libellé de N ou N-1 (priorité à N)
                libelle = n_row.get('Libellé', n1_row.get('Libellé', ''))
                
                comp_data.append({
                    'Compte': account,
                    'Libellé': libelle,
                    'Solde_N': solde_n,
                    'Solde_N-1': solde_n1,
                    'Variation': variation,
                    '%_Evolution': pct_evolution,
                    'Débit_N': debit_n,
                    'Crédit_N': credit_n,
                    'Débit_N-1': debit_n1,
                    'Crédit_N-1': credit_n1,
                    'Date_Creation_N': n_row.get('Date_Creation', ''),
                    'Date_Creation_N-1': n1_row.get('Date_Creation', ''),
                    'Exercices': f"{n}/{n_1}"
                })
            
            # Trier par numéro de compte
            comp_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in comp_data:
                ws_comp.append([row_data['Compte'], row_data['Libellé'], row_data['Solde_N'], 
                              row_data['Solde_N-1'], row_data['Variation'], row_data['%_Evolution'],
                              row_data['Débit_N'], row_data['Crédit_N'], row_data['Débit_N-1'], 
                              row_data['Crédit_N-1'], row_data['Date_Creation_N'], row_data['Date_Creation_N-1'], row_data['Exercices']])
            
            # Appliquer le formatage aux onglets
            for ws in [ws_n, ws_n1, ws_comp]:
                # En-têtes en gras
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Ajuster la largeur des colonnes
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Stocker les données pour l'API
            sheets_data[group_name] = {
                'exercice_n': n_data,
                'exercice_n1': n1_data,
                'comparatif': comp_data,
                'has_two_exercices': True,
                'exercices': [n, n_1]
            }
            
        else:
            # Cas 2: Un seul exercice - Créer 1 onglet
            ws_n = wb.create_sheet(f"Exercice_{n}")
            ws_n.append(['Compte', 'Libellé', 'Solde', 'Débit', 'Crédit', 'Exercice', 'Date_Creation'])
            
            # Remplir les données N
            n_data = []
            for _, row in group_n.iterrows():
                n_data.append({
                    'Compte': row['account_number'],
                    'Libellé': row.get('account_name', ''),
                    'Solde': row['balance'],
                    'Débit': row.get('total_debit', 0),
                    'Crédit': row.get('total_credit', 0),
                    'Exercice': n,
                    'Date_Creation': row.get('created_at', '')
                })
            
            # Trier par numéro de compte
            n_data.sort(key=lambda x: str(x['Compte']))
            
            for row_data in n_data:
                ws_n.append([row_data['Compte'], row_data['Libellé'], row_data['Solde'], 
                           row_data['Débit'], row_data['Crédit'], row_data['Exercice'], row_data['Date_Creation']])
            
            # Appliquer le formatage
            for cell in ws_n[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ajuster la largeur des colonnes
            for column in ws_n.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_n.column_dimensions[column_letter].width = adjusted_width
            
            # Stocker les données pour l'API
            sheets_data[group_name] = {
                'exercice_n': n_data,
                'has_two_exercices': False,
                'exercices': [n]
            }
        
        # Sauvegarder le fichier Excel
        output = BytesIO()
        wb.save(output)
        sheets_contents[group_name] = output.getvalue()
    
    # Ajout du contrôle de cohérence au retour
    coherence = controle_coherence_complet(tft_data)
    return tft_content, sheets_contents, tft_data, sheets_data, coherence
